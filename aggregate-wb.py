import os
from openpyxl import Workbook, load_workbook

wb_folder = "workbooks"
output_workbook = "aggregate_workbook.xlsx"


def main():
    struct = get_structure()

    aggregate = Workbook()
    aggregate.remove(aggregate.active)
    aggregate_sheets = {}
    for sheetname in struct.keys():
        sheet = aggregate.create_sheet(sheetname)
        sheet.append(struct[sheetname])
        aggregate_sheets.update({sheetname: sheet})

    for filename in os.listdir(wb_folder):
        wb = load_workbook(filename=wb_folder + "/" + filename)
        for sheetname in struct.keys():
            write_sheet = aggregate_sheets[sheetname]
            read_sheet = wb[sheetname]
            for row in read_sheet.iter_rows(min_row=2):
                row_list = []
                for cell in row:
                    row_list.append(cell.value)
                write_sheet.append(row_list)
    aggregate.save(output_workbook)


def get_structure():
    out = {}
    wb = load_workbook(filename=wb_folder + "/" + os.listdir(wb_folder)[0])
    for sheetname in wb.sheetnames:
        row = []
        for cell in wb[sheetname][1]:
            row.append(cell.value)
        out.update({sheetname: row})
    return out


if __name__ == '__main__':
    main()
